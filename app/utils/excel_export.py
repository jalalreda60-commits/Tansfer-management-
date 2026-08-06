"""
Generic Excel export helper: dumps a list of dict rows (or a QTableWidget)
to a styled .xlsx workbook.
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.config import EXPORTS_DIR

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_rows_to_excel(headers: list[str], rows: list[list], sheet_title: str,
                          filename: str | None = None) -> str:
    """Write headers+rows to an .xlsx file under EXPORTS_DIR, returns the path."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Export"

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    # Auto width
    for col_idx, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows] or [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 60)

    ws.freeze_panes = "A2"

    if not filename:
        stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{sheet_title.replace(' ', '_')}_{stamp}.xlsx"
    path = os.path.join(EXPORTS_DIR, filename)
    wb.save(path)
    return path


def export_table_widget_to_excel(table_widget, sheet_title: str) -> str:
    """Export a QTableWidget's visible contents to Excel."""
    headers = [table_widget.horizontalHeaderItem(c).text() if table_widget.horizontalHeaderItem(c) else ""
               for c in range(table_widget.columnCount())]
    rows = []
    for r in range(table_widget.rowCount()):
        row = []
        for c in range(table_widget.columnCount()):
            item = table_widget.item(r, c)
            row.append(item.text() if item else "")
        rows.append(row)
    return export_rows_to_excel(headers, rows, sheet_title)
